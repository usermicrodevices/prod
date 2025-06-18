import base64, logging, os, re, sys, time
from decimal import Decimal
from io import BytesIO, StringIO
from subprocess import Popen, PIPE
from datetime import datetime, timedelta
try:
    from zoneinfo import available_timezones, ZoneInfo
except:
    from backports.zoneinfo import available_timezones, ZoneInfo

from django.utils import timezone as django_timezone
from django.utils.translation import gettext as _
from django.utils.html import format_html, format_html_join, html_safe
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.contrib import admin, messages
from django import forms
from django.http import StreamingHttpResponse, FileResponse, HttpResponseRedirect
from django.contrib.auth.forms import ReadOnlyPasswordHashField
from django.db.models import F, Q, Min, Max, Sum, When, Value, Count, IntegerField, TextField, CharField, OuterRef, Subquery
from django.db.models.query import QuerySet
from django.db import connections
from django.contrib.admin.models import LogEntry
from django.contrib.admin.widgets import AutocompleteSelect
from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.shortcuts import render
from django.views.generic.edit import FormView
from django.core.cache import caches
from django.conf import settings
from django.apps import apps as django_apps
from django.template import Context, Template

from .models import PrintTemplates, Unit, Currency, Country, Region, City, Tax, CompanyType, Company, SalePoint, Manufacturer, ProductModel, BarCode, QrCode, DocType, ProductGroup, Product, Customer, ProductImage
from users.models import User


def get_model(app_model):
    app_name, model_name = app_model.split('.')
    return django_apps.get_app_config(app_name).get_model(model_name)

def get_queryset_by_company(super, request):
    qs = super.get_queryset(request)
    user = request.user
    if user.is_superuser:
        return qs
    return qs.filter(Q(company__in=user.companies.all()) | Q(company_id__in=user.sale_points.values_list('company_id', flat=True))).distinct()

def get_queryset_by_sale_point(super, request):
    qs = super.get_queryset(request)
    user = request.user
    if user.is_superuser:
        return qs
    return qs.filter(Q(sale_point__company__in=user.companies.all()) | Q(sale_point__in=user.sale_points.all())).distinct()


class RawImage:
    _id = 1
    format = "png"
    path = f"/xl/media/image{_id}.{format}"
    anchor = "A1"
    width, height = 0, 0
    def __init__(self, img):
        self.ref = img
    def _data(self):
        return self.ref

def raw_find_images(archive, path):
    from openpyxl.xml.constants import IMAGE_NS
    from openpyxl.xml.functions import fromstring
    from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
    from openpyxl.packaging.relationship import get_rels_path, get_dependents
    images = []
    src = archive.read(path)
    tree = fromstring(src)
    try:
        drawing = SpreadsheetDrawing.from_tree(tree)
    except TypeError:
        return [], images
    rels_path = get_rels_path(path)
    deps = []
    if rels_path in archive.namelist():
        deps = get_dependents(archive, rels_path)
    for rel in drawing._blip_rels:
        dep = deps.get(rel.embed)
        if dep.Type == IMAGE_NS:
            try:
                image = RawImage(archive.read(dep.target))
            except OSError:
                continue
            image.anchor = rel.anchor
            images.append(image)
    return [], images


class DropDownFilter(admin.SimpleListFilter):
    template = 'dropdown_filter_from_memory.html'


class UploadFileForm(forms.Form):
    _selected_action = forms.CharField(widget=forms.MultipleHiddenInput)
    file = forms.FileField(widget=forms.ClearableFileInput(attrs={'allow_multiple_selected': True}))


class CountryFilter(DropDownFilter):
    title = _('Country')
    parameter_name = 'country'

    def lookups(self, request, model_admin):
        res = []
        queryset = Country.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(country=self.value())


class RegionFilter(DropDownFilter):
    title = _('Region')
    parameter_name = 'region'

    def lookups(self, request, model_admin):
        res = []
        queryset = Region.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(region=self.value())


class CityFilter(DropDownFilter):
    title = _('City')
    parameter_name = 'city'

    def lookups(self, request, model_admin):
        res = []
        queryset = City.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(city=self.value())


class UnitFilter(DropDownFilter):
    title = _('Unit')
    parameter_name = 'unit'

    def lookups(self, request, model_admin):
        res = []
        queryset = Unit.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(unit=self.value())


class CurrencyFilter(DropDownFilter):
    title = _('Currency')
    parameter_name = 'currency'

    def lookups(self, request, model_admin):
        res = []
        queryset = Currency.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(currency=self.value())


class TaxFilter(DropDownFilter):
    title = _('Tax')
    parameter_name = 'tax'

    def lookups(self, request, model_admin):
        res = []
        queryset = Tax.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(tax=self.value())


class ProductModelFilter(DropDownFilter):
    title = _('ProductModel')
    parameter_name = 'model'

    def lookups(self, request, model_admin):
        res = []
        queryset = ProductModel.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(model=self.value())


class CompanyTypeFilter(DropDownFilter):
    title = _('Company Type')
    parameter_name = 'type'

    def lookups(self, request, model_admin):
        user = request.user
        res = []
        queryset = CompanyType.objects.only('id', 'name')
        if not user.is_superuser:
            queryset = queryset.filter(Q(type__company__in=user.companies.all()) | Q(type__company__id__in=user.sale_points.values_list('company_id', flat=True)))
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(type=self.value())


class CompanyFilter(DropDownFilter):
    title = _('Company')
    parameter_name = 'company'

    def lookups(self, request, model_admin):
        user = request.user
        res = []
        queryset = Company.objects.only('id', 'name')
        if not user.is_superuser:
            queryset = queryset.filter(Q(pk__in=user.companies.all()) | Q(id__in=user.sale_points.values_list('company_id', flat=True)))
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(company=self.value())


class TypeCompanyFilter(CompanyFilter):

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(type__company=self.value())


class SalePointCompanyFilter(CompanyFilter):

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(sale_point__company=self.value())


class SalePointFilter(DropDownFilter):
    title = _('Sale Point')
    parameter_name = 'sale_point'

    def lookups(self, request, model_admin):
        user = request.user
        res = []
        queryset = SalePoint.objects.only('id', 'name')
        if not user.is_superuser:
            queryset = queryset.filter(Q(company__in=user.companies.all()) | Q(id__in=user.sale_points.values_list('id', flat=True)))
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(sale_point=self.value())


class BarCodeFilter(DropDownFilter):
    title = _('BarCode')
    parameter_name = 'barcodes__in'

    def lookups(self, request, model_admin):
        res = []
        queryset = BarCode.objects.only('id')
        for it in queryset:
            res.append((it.id, it.id))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(barcodes__in=[self.value()])


class QrCodeFilter(DropDownFilter):
    title = _('QrCode')
    parameter_name = 'qrcodes__in'

    def lookups(self, request, model_admin):
        res = []
        queryset = QrCode.objects.only('id')
        for it in queryset:
            res.append((it.id, it.id))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(qrcodes__in=[self.value()])


class ProductGroupFilter(DropDownFilter):
    title = _('Product Group')
    parameter_name = 'group'

    def lookups(self, request, model_admin):
        res = []
        queryset = ProductGroup.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(group=self.value())


class ProductManufacturerFilter(DropDownFilter):
    title = _('Product Manufacturer')
    parameter_name = 'model__manufacturer'

    def lookups(self, request, model_admin):
        res = []
        queryset = Manufacturer.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(model__manufacturer=self.value())


class ProductFilter(DropDownFilter):
    title = _('Product')
    parameter_name = 'product'

    def lookups(self, request, model_admin):
        res = []
        queryset = Product.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(product=self.value())


class DocTypeFilter(DropDownFilter):
    title = _('Doc Type')
    parameter_name = 'type'

    def lookups(self, request, model_admin):
        res = []
        queryset = DocType.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(type=self.value())


class CustomerFilter(DropDownFilter):
    title = _('Customer')
    parameter_name = 'customer'

    def lookups(self, request, model_admin):
        res = []
        queryset = Customer.objects.only('id', 'name')
        for it in queryset:
            res.append((it.id, it.name))
        return res

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        else:
            return queryset.filter(customer=self.value())


class CustomModelAdmin(admin.ModelAdmin):

    def logi(self, *args):
        msg = f'ℹ️{self.__class__.__name__}.{sys._getframe().f_back.f_code.co_name}'
        for arg in args:
            msg += f'::{arg}'
        logging.info(msg)

    def logw(self, *args):
        msg = f'⚠{self.__class__.__name__}.{sys._getframe().f_back.f_code.co_name}'
        for arg in args:
            msg += f'::{arg}'
        logging.warning(msg)

    def logd(self, *args):
        msg = f'⚠{self.__class__.__name__}.{sys._getframe().f_back.f_code.co_name}'
        for arg in args:
            msg += f'::{arg}'
        logging.debug(msg)

    def loge(self, err, *args):
        msg = f'🆘{self.__class__.__name__}.{err.__traceback__.tb_frame.f_code.co_name}'
        for arg in args:
            msg += f'::{arg}'
        msg += f'::{err}::LINE={err.__traceback__.tb_lineno}'
        logging.error(msg)

    def noselect_actions(self, request, auto_selectable_actions=[]):
        if 'action' in request.POST and request.POST['action'] in auto_selectable_actions:
            if not request.POST.getlist(ACTION_CHECKBOX_NAME):
                post = request.POST.copy()
                post.update({ACTION_CHECKBOX_NAME:'0'})
                request._set_post(post)
        return request

    def worksheet_cell_write(self, worksheet, row, col, value, type_value = None, fmt = None):
        func_write = worksheet.write
        if type_value == 'as_number':
            func_write = worksheet.write_number
        elif type_value == 'as_datetime':
            func_write = worksheet.write_datetime
        elif type_value == 'as_base64':
            try:
                worksheet.embed_image(row, col, fmt, {'image_data': value})
            except Exception as e:
                self.loge(e, row, col, fmt)
            return col + 1
        try:
            if fmt:
                func_write(row, col, value, fmt)
            else:
                func_write(row, col, value)
        except Exception as e:
            try:
                if fmt:
                    func_write(row, col, repr(value), fmt)
                else:
                    func_write(row, col, repr(value))
            except Exception as e:
                self.loge(e, row, col)
        return col + 1

    def queryset_to_xls(self, queryset, fields={}, exclude_fields=['id']):
        import xlsxwriter
        output = None
        if queryset.count():
            field_names = list(fields.keys())
            if not field_names:
                for field in queryset.model._meta.get_fields():
                    if field.name and field.name not in exclude_fields:
                        field_names.append(field.name)
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet()
            cell_format_bold = workbook.add_format({'align':'center', 'valign':'vcenter', 'bold':True})
            cell_format_left = workbook.add_format({'align':'left', 'valign':'vcenter'})
            col = 0
            for field_name in field_names:
                field_title = field_name
                if field_name in fields:
                    width = fields[field_name].get('width', None)
                    if width is not None:
                        worksheet.set_column(col, col, width)
                    field_title = fields[field_name].get('title', field_title)
                col = self.worksheet_cell_write(worksheet, 0, col, _(field_title), fmt=cell_format_bold)
            row = 1
            for item in queryset:
                worksheet.set_row(row, None, cell_format_left)
                col = 0
                if settings.DEBUG:
                    record = {}
                for field_name in field_names:
                    if not hasattr(item, field_name):
                        col += 1
                        continue
                    else:
                        if not field_name:
                            col += 1
                            continue
                    try:
                        value = getattr(item, field_name)
                    except AttributeError as e:
                        col += 1
                        self.loge(e)
                    except Exception as e:
                        col += 1
                        self.loge(e)
                    else:
                        if settings.DEBUG:
                            record[field_name] = value
                        if not value:
                            col += 1
                        else:
                            format_value = None
                            tvalue = None
                            if isinstance(value, datetime):
                                value = f'{value.strftime("%Y.%m.%d %H:%M:%S")}'
                            elif isinstance(value, (int, float, Decimal)):
                                tvalue = 'as_number'
                            elif not isinstance(value, str):
                                value = f'{value}'
                            elif isinstance(value, str) and 'data:image/' in value and ';base64,' in value:
                                #with open(f'{row}{col}.png', 'wb') as fpng:
                                    #fpng.write(base64.decodebytes(value.split(';base64,')[1].encode()))
                                value = BytesIO(base64.decodebytes(value.split(';base64,')[1].encode()))
                                tvalue = 'as_base64'
                                format_value = f'{row}{col}.png'
                            col = self.worksheet_cell_write(worksheet, row, col, value, tvalue, format_value)
                if settings.DEBUG:
                    self.logd('ROW', row, record)
                row += 1
            workbook.close()
            output.seek(0)
        return output


class PrintTemplatesAdmin(CustomModelAdmin):
    list_display = ('id', 'alias', 'content', 'extinfo')
    list_display_links = ('id', 'alias')
    search_fields = ('id', 'alias', 'content', 'extinfo')
admin.site.register(PrintTemplates, PrintTemplatesAdmin)


class UnitAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'label')
    list_editable = ('name', 'label')
    search_fields = ('id', 'name', 'label')
admin.site.register(Unit, UnitAdmin)


class CurrencyAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'alias')
    list_display_links = ('name',)
    search_fields = ('id', 'name', 'alias')
admin.site.register(Currency, CurrencyAdmin)


class TaxAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'alias', 'value')
    list_display_links = ('name',)
    search_fields = ('id', 'name', 'alias', 'value')
admin.site.register(Tax, TaxAdmin)


class CountryAdmin(CustomModelAdmin):
    list_display = ('id', 'name')
    list_display_links = ['name']
    search_fields = ['id', 'name']
admin.site.register(Country, CountryAdmin)


class RegionAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'country')
    list_display_links = ['name']
    search_fields = ['id', 'name', 'country__name']
    list_select_related = ['country']
    list_filter = (CountryFilter,)
    autocomplete_fields = ['country']
    raw_id_fields = ['country']
admin.site.register(Region, RegionAdmin)


class CityAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'region')
    list_display_links = ['id', 'name']
    search_fields = ['id', 'name', 'region__name', 'region__country__name']
    list_select_related = ['region__country']
    list_filter = (CountryFilter, RegionFilter)
    autocomplete_fields = ['region']
    raw_id_fields = ['region']
admin.site.register(City, CityAdmin)


class ManufacturerAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'city', 'get_region', 'get_country')
    list_display_links = ('id', 'name')
    search_fields = ('id', 'name', 'city__name', 'city__region__name', 'city__region__country__name')
    list_filter = (CountryFilter, RegionFilter, CityFilter)
    autocomplete_fields = ['city']

    def get_region(self, obj):
        try:
            return obj.city.region.name
        except:
            return ''
    get_region.short_description = _('region')

    def get_country(self, obj):
        try:
            return obj.city.region.country.name
        except:
            return ''
    get_country.short_description = _('country')

admin.site.register(Manufacturer, ManufacturerAdmin)


class ProductModelAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'manufacturer')
    list_display_links = ('id', 'name')
    search_fields = ('id', 'name', 'manufacturer__name')
admin.site.register(ProductModel, ProductModelAdmin)


class CompanyTypeAdmin(CustomModelAdmin):
    list_display = ('id', 'name')
    list_display_links = ('id', 'name')
    search_fields = ('id', 'name')
admin.site.register(CompanyType, CompanyTypeAdmin)


class DocTypeAdmin(CustomModelAdmin):
    list_display = ('id', 'alias', 'name', 'income', 'auto_register', 'description', 'get_count')
    list_display_links = ('id', 'alias', 'name')
    search_fields = ('id', 'alias', 'name', 'description')

    def get_count(self, obj):
        return format_html('<a href="{}/core/doc/?type_id={}" target="_blank">{}</a>', settings.ADMIN_PATH_PREFIX, obj.id, get_model('core.Doc').objects.filter(type_id=obj.id).count())
    get_count.short_description = _('count')

admin.site.register(DocType, DocTypeAdmin)


class CompanyAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'created_date', 'type', 'currency', 'extinfo')
    list_display_links = ['id', 'name']
    search_fields = ['id', 'name', 'created_date', 'type__name',  'currency__name', 'currency__alias']
    list_select_related = ['type']
    autocomplete_fields = ['type']
    list_filter = [CompanyTypeFilter,  CurrencyFilter]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        user = request.user
        if not user.is_superuser:
            qs = qs.filter(Q(pk__in=user.companies.values_list('id', flat=True)) | Q(pk__in=user.sale_points.values_list('company_id', flat=True))).distinct()
        return qs

admin.site.register(Company, CompanyAdmin)


class SalePointAdmin(CustomModelAdmin):
    save_on_top = True
    list_display = ('id', 'name', 'get_company', 'created_date', 'address', 'map_point', 'city')
    list_display_links = ['id', 'name']
    search_fields = ['name', 'company__name', 'created_date', 'address', 'map_point']
    list_select_related = ['company']
    list_filter = [CompanyFilter, CityFilter]
    autocomplete_fields = ['company']
    raw_id_fields = ['city']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        user = request.user
        if user.is_superuser:
            return qs
        return qs.filter(Q(company__in=user.companies.all()) | Q(pk__in=user.sale_points.values_list('id', flat=True))).distinct()

    def get_company(self, obj):
        o = obj.company
        if not o:
            return ''
        m = o._meta
        return format_html('''<p><font color="green" face="Verdana, Geneva, sans-serif"><a href="{}" target="_blank">{}</a></font></p>''', reverse('admin:{}_{}_change'.format(m.app_label, m.model_name), args=(o.id,)), o.name)
    get_company.short_description = _('Company')
    get_company.admin_order_field = 'company'

admin.site.register(SalePoint, SalePointAdmin)


class BarCodeAdmin(CustomModelAdmin):
    list_display = ('id', 'get_products')
    list_display_links = ['id']
    search_fields = ['id']
    actions = ('fix_code', 'delete_ghosts')

    def changelist_view(self, request, extra_context=None):
        request = self.noselect_actions(request, ['delete_ghosts'])
        return super().changelist_view(request, extra_context)

    def get_products(self, obj):
        try:
            idxs = Product.objects.filter(barcodes__id=obj.id).annotate(admin_path_prefix=Value(settings.ADMIN_PATH_PREFIX, CharField())).values_list('admin_path_prefix', 'id', 'name')
        except Exception as e:
            self.loge(e)
            return ''
        else:
            if not idxs:
                return ''
            content = format_html_join('\n', '<p><font color="green" face="Verdana, Geneva, sans-serif"><a href="{0}/refs/product/?id={1}" target="_blank">{2}</a></font></p>', idxs)
            return format_html('<details><summary>{}</summary>{}</details>', idxs[0][2], content)
    get_products.short_description = _('Products')

    def fix_code(self, request, queryset):
        from .models import ean13
        replaced = 0
        for bcode in queryset:
            e13 = ean13(bcode.id)
            if bcode.id != e13:
                products = []
                for p in Product.objects.filter(barcodes__id=bcode.id):
                    products.append(p)
                    p.barcodes.remove(bcode)
                bcode.id = e13
                try:
                    bcode.save()
                except Exception as e:
                    self.loge(e)
                else:
                    replaced += 1
                    for p in products:
                        p.barcodes.add(bcode)
        self.message_user(request, f"{_(f'fixed')} {replaced}", messages.SUCCESS)
    fix_code.short_description = f'✨{_("fix values")}🖋'

    def delete_ghosts(self, request, queryset):
        if not request.user.is_superuser and not request.user.has_perm('refs.delete_barcode'):
            self.message_user(request, _('not permission for this operation'), messages.WARNING)
        else:
            selected = queryset.count()
            if not selected:
                queryset = self.model.objects.all()
                selected = queryset.count()
            if settings.DEBUG:
                self.logd('SELECTED-COUNT', selected)
            if not selected:
                self.message_user(request, _('please select barcodes'), messages.WARNING)
            else:
                try:
                    count_delete, del_refs = queryset.filter(product__isnull=True).delete()
                except Exception as e:
                    self.loge(e)
                    self.message_user(request, f'{e}', messages.ERROR)
                else:
                    self.message_user(request, f'{_("delete barcodes")} {count_delete} {del_refs}', messages.SUCCESS)
    delete_ghosts.short_description = f'𝔻𝓓🇩{_("delete all without products")}𝕯Ⓓ𝐃'

admin.site.register(BarCode, BarCodeAdmin)


class QrCodeAdmin(CustomModelAdmin):
    list_display = ['id']
    list_display_links = ['id']
    search_fields = ['id']
admin.site.register(QrCode, QrCodeAdmin)


class ProductGroupAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'parent', 'alias', 'description', 'extinfo')
    list_display_links = ('id', 'name', 'alias')
    search_fields = ('id', 'name', 'alias', 'description', 'extinfo')
    list_filter = ('parent',)
admin.site.register(ProductGroup, ProductGroupAdmin)


@html_safe
class JSProduct:
    def __str__(self):
        return r'''<script>'use strict';
document.addEventListener("DOMContentLoaded", function(event) {
setTimeout(() => {
    const elements = document.getElementsByClassName('product_count');
    for(var i = 0; i < elements.length; i++)
    {
        var column = elements[i];
        if(column && column.style)
        {
            var row = column.parentNode.parentNode;
            if(row)
            {
                row.style.cssText = column.style.cssText;
                const div = row.querySelectorAll("div.related-widget-wrapper[data-model-ref='tax']")[0];
                if(div)
                {
                    const spans = div.getElementsByClassName('select2 select2-container select2-container--admin-autocomplete');
                    Array.from(spans).forEach(span => {span.style.width = "100px";});
                }
            }
        }
    }
});});</script>'''


class ProductAdmin(CustomModelAdmin):
    __objs__ = {}
    user = None
    list_display = ['id', 'article', 'name', 'get_barcodes', 'get_qrcodes', 'get_price', 'get_count', 'get_sum', 'get_prod_model', 'get_group', 'get_thumbnail', 'extinfo', 'tax']
    list_display_links = ('id', 'article', 'name')
    search_fields = ('id', 'name', 'article', 'extinfo', 'barcodes__id', 'qrcodes__id', 'group__name', 'model__name', 'model__manufacturer__name')
    list_select_related = ('tax', 'model', 'group')
    raw_id_fields = ['barcodes', 'qrcodes', 'images']
    list_editable = ['tax']
    autocomplete_fields = ['model', 'tax', 'group']
    list_filter = (ProductGroupFilter, ProductManufacturerFilter, ProductModelFilter, TaxFilter)
    actions = ('order_from_selected_items',
        'from_xls_with_check',
        'from_xls',
        'to_xls',
        'price_to_xls',
        'barcode_to_svg',
        'qr_to_svg',
        'fix_barcodes',
        'copy_name_to_ext_label',
        'copy_unit',
        'copy_cost',
        'copy_price',
        'copy_cost_price',
        'thumbnails_from_xls',
        'thumbnails_to_xls',
        'thumbnail_from_first_image',
        'thumbnail_clear',
        'reset_cached')

    class Media:
        #css = {'screen': ['admin/css/vendor/select2/select2.css', '/static/admin/css/autocomplete.css']}
        js = [JSProduct()]#('admin/js/vendor/jquery/jquery.js', , 'admin/js/vendor/select2/select2.full.js', 'admin/js/autocomplete.js')

    def check_cost_permission(self):
        if self.user.is_superuser:
            return True
        model_name = self.__class__.__name__.replace('Admin', '')
        if get_model('users.RoleField').objects.filter(role=self.user.role, role_model__app='core', role_model__model=model_name, read=True, value='cost').exists():
            return True
        return False

    def changelist_view(self, request, extra_context=None):
        self.user = request.user
        if settings.DEBUG:
            self.logd('CURRENT USER', self.user)
        if self.check_cost_permission():
            if 'get_cost' not in self.list_display:
                self.list_display.insert(self.list_display.index('get_price'), 'get_cost')
        elif 'get_cost' in self.list_display:
            self.list_display.remove('get_cost')
        request = self.noselect_actions(request, ['from_xls_with_check', 'from_xls', 'thumbnails_from_xls', 'reset_cached'])
        template_response = super().changelist_view(request, extra_context)
        #TODO change row styles on template render
        #self.logi(template_response.context_data, template_response.template_name)
        #template = template_response.resolve_template(template_response.template_name)
        #context = template_response.resolve_context(template_response.context_data)
        #self.logi(template.template.source)
        #cl = self.get_changelist_instance(request)
        #self.logi(cl.result_list)
        return template_response

    def get_form(self, request, obj=None, **kwargs):
        self.user = request.user
        if settings.DEBUG:
            self.logd('CURRENT USER', self.user)
        if self.check_cost_permission():
            if self.exclude:
                self.exclude = ()
        else:
            self.exclude = ('cost', 'get_cost')
        form = super().get_form(request, obj, **kwargs)
        form.current_user = self.user
        if not obj:
            last_prod = Product.objects.order_by('article').last()
            if last_prod and last_prod.article:
                form.base_fields['article'].initial = ''.join([f'{int(it)+1}' if it.isdigit() else it for it in re.split(r'\W+', last_prod.article)])
        return form

    def get_last_reg(self, obj):
        if obj.id in self.__objs__ and 'lreg' in self.__objs__[obj.id]:
            return self.__objs__[obj.id]['lreg']
        else:
            if obj.id not in self.__objs__:
                self.__objs__[obj.id] = {}
            try:
                self.__objs__[obj.id]['lreg'] = get_model('core.Register').objects.filter(rec__product_id=obj.id).order_by('-rec__doc__registered_at').select_related('rec').first()
            except Exception as e:
                self.loge(e)
            else:
                return self.__objs__[obj.id]['lreg']

    def get_price_value(self, obj):
        if settings.BEHAVIOR_PRICE.get('select_from_register', False):
            last_reg = self.get_last_reg(obj)
            if last_reg:
                return last_reg.rec.price
        return obj.price

    def get_cost(self, obj):
        if not self.user.is_superuser and not get_model('users.RoleField').objects.filter(role=self.user.role, role_model__app='core', role_model__model=obj.__class__.__name__, read=True, value='cost').exists():
            return ''
        value = obj.cost
        if settings.BEHAVIOR_COST.get('select_from_register', False):
            last_reg = self.get_last_reg(obj)
            if last_reg:
                value = last_reg.rec.cost
        return format_html('<font color="green" face="Verdana, Geneva, sans-serif">{} {}</font>', value, obj.currency.name if obj.currency else '')
    get_cost.short_description = _('cost')
    get_cost.admin_order_field = 'cost'

    def get_price(self, obj):
        return format_html('<font color="green" face="Verdana, Geneva, sans-serif">{} {}</font>', self.get_price_value(obj).quantize(Decimal('0.00')), obj.currency.name if obj.currency else '')
    get_price.short_description = _('price')
    get_price.admin_order_field = 'price'

    def get_barcodes(self, obj):
        try:
            idxs = BarCode.objects.filter(product__in=[obj]).annotate(admin_path_prefix=Value(settings.ADMIN_PATH_PREFIX, CharField())).values_list('admin_path_prefix', 'id')
        except Exception as e:
            return ''
        else:
            if not idxs:
                return ''
            content = format_html_join('\n', '<p><font color="green" face="Verdana, Geneva, sans-serif"><a href="{0}/refs/barcode/?id={1}" target="_blank">{1}</a></font></p>', idxs)
            return format_html('<details><summary>{}</summary>{}</details>', idxs[0][1], content)
    get_barcodes.short_description = _('Bar Codes')

    def get_qrcodes(self, obj):
        try:
            idxs = QrCode.objects.filter(product__in=[obj]).annotate(admin_path_prefix=Value(settings.ADMIN_PATH_PREFIX, CharField())).values_list('admin_path_prefix', 'id')
        except Exception as e:
            return ''
        else:
            if not idxs:
                return ''
            content = format_html_join('\n', '<p><font color="green" face="Verdana, Geneva, sans-serif"><a href="{0}/refs/qrcode/?id={1}" target="_blank">{1}</a></font></p>', idxs)
            return format_html('<details><summary>{}</summary>{}</details>', idxs[0][1], content)
    get_qrcodes.short_description = _('Qr Codes')

    def refresh_count_from_reg(self, obj):
        if obj.id not in self.__objs__:
            self.__objs__[obj.id] = {}
        SumIncome=Sum('rec__count', filter=Q(rec__doc__type__income=True), default=0)
        SumExpense=Sum('rec__count', filter=Q(rec__doc__type__income=False), default=0)
        try:
            self.__objs__[obj.id]['count'] = get_model('core.Register').objects.filter(rec__product_id=obj.id).aggregate(count=SumIncome-SumExpense)['count']
        except Exception as e:
            self.loge(e)
        else:
            if settings.DEBUG:
                self.logi('FROM BASE', self.__objs__[obj.id]['count'])
            return self.__objs__[obj.id]['count']
        return 0

    def get_count_from_reg(self, obj):
        if obj.id in self.__objs__ and 'count' in self.__objs__[obj.id]:
            if settings.DEBUG:
                self.logi('FROM BUFFER', self.__objs__[obj.id]['count'])
            return self.__objs__[obj.id]['count']
        return self.refresh_count_from_reg(obj)

    def get_min_styles_reversed(self, grp, min_styles = {}):
        min_styles = grp.extinfo.get('min_styles', {})
        if grp.parent:
            min_styles |= self.get_min_styles_reversed(grp.parent, min_styles)
        return min_styles

    def get_min_styles(self, grp, min_styles = {}):
        if grp.parent:
            min_styles |= self.get_min_styles(grp.parent, min_styles)
        return min_styles | grp.extinfo.get('min_styles', {})

    def get_count(self, obj):
        count = self.get_count_from_reg(obj)
        color = 'color:green;' if count > 0 else 'color:red;'
        if obj.group:
            min_styles = self.get_min_styles(obj.group)
            if min_styles:
                #for k in sorted(min_styles.keys(), reverse=True):
                for k in min_styles.keys():
                    if count <= int(k):
                        color = min_styles[k]
                        break
        return format_html('<p class="product_count" id={} style="{}"><a href="{}/core/register/?rec__product={}" target="_blank" style="{}">{} {}</a></p>', obj.id, color, settings.ADMIN_PATH_PREFIX, obj.id, color, count, obj.unit.label)
    get_count.short_description = _('Count')

    def get_sum(self, obj):
        count = self.get_count_from_reg(obj)
        price = self.get_price_value(obj)
        if count > 0 and price > 0:
            #return f'{count * price:.2f}'
            return format_html('<font color="green" face="Verdana, Geneva, sans-serif">{} {}</font>', (price*count).quantize(Decimal('0.00')), obj.currency.name if obj.currency else '')
        return '0'
    get_sum.short_description = _('sum')

    def get_tax(self, obj):
        o = obj.tax
        if not o:
            return ''
        m = o._meta
        return format_html('<a href="{}?id={}" target="_blank">{}</a>', reverse(f'admin:{m.app_label}_{m.model_name}_changelist'), o.id, o.name)
    get_tax.short_description = _('Tax')
    get_tax.admin_order_field = 'tax'

    def get_prod_model(self, obj):
        o = obj.model
        if not o:
            return ''
        m = o._meta
        return format_html('<a href="{}?id={}" target="_blank">{} {}</a>', reverse(f'admin:{m.app_label}_{m.model_name}_changelist'), o.id, o.manufacturer.name, o.name)
    get_prod_model.short_description = _('Model')
    get_prod_model.admin_order_field = 'model'

    def get_group(self, obj):
        o = obj.group
        if not o:
            return ''
        m = o._meta
        return format_html('<a href="{}?id={}" target="_blank">{}</a>', reverse(f'admin:{m.app_label}_{m.model_name}_changelist'), o.id, o.name)
    get_group.short_description = _('Product Group')
    get_group.admin_order_field = 'group'

    def reset_cached(self, request, queryset):
        self.__objs__ = {}
    reset_cached.short_description = f'↻💦{_("reset cached values")}'

    def get_thumbnail(self, obj):
        if obj.thumbnail:
            return format_html('<img src="{}" width="32" height="32">', obj.thumbnail)
        else:
            return format_html('<img src="{}" width="32" height="32">', settings.DEFAUL_IMAGE_THUMBNAIL)
    get_thumbnail.short_description = _('thumbnail')

    def thumbnail_clear(self, request, queryset):
        updated_products, errs = [], ''
        for product in queryset:
            if product.thumbnail:
                product.thumbnail = None
                updated_products.append(product)
        if updated_products:
            try:
                Product.objects.bulk_update(updated_products, ['thumbnail'])
            except Exception as e:
                self.loge(e)
                errs = f'; {e}'
                updated_products = []
        self.message_user(request, f'📄 {_("cleared thumbnails")} {len(updated_products)}{errs}📄', messages.SUCCESS)
    thumbnail_clear.short_description = f'📄♻️{_("clear thumbnail")}🚮📄'

    def thumbnail_from_first_image(self, request, queryset):
        result_count = 0
        for item in queryset:
            img = item.images.order_by('id').first()
            if img and img.file and os.path.isfile(img.file.path):
                self.logi(img.file.path)
                try:
                    convert = Popen(['convert', f'{img.file.path}', '-resize', '256x256', '-'], stdin=PIPE, stdout=PIPE, stderr=PIPE)
                except Exception as e:
                    self.loge(e)
                else:
                    stdout_data, stderr_data = convert.communicate(b'base64')
                    if stderr_data:
                        self.logw(stderr_data, convert.args)
                    elif stdout_data:
                        try:
                            item.thumbnail = f'data:image/{img.file.path.split(".")[-1]};base64,{base64.b64encode(stdout_data).decode()}'
                        except Exception as e:
                            self.loge(e, img.file.path)
                        else:
                            try:
                                item.save()
                            except Exception as e:
                                self.loge(e, item)
        self.message_user(request, f'📄 {_("filled thumbnails")} {result_count} 📄', messages.SUCCESS)
    thumbnail_from_first_image.short_description = f'📄{_("thumbnail from first image")}📄'

    def thumbnails_from_xls(self, request, queryset):
        from openpyxl import reader
        from openpyxl.drawing.image import PILImage
        form = None
        if 'apply' in request.POST:
            self.logi('💡', request.FILES)
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                msg_err = ''
                updated_products = []
                file = form.cleaned_data['file']
                if file:
                    # memory optimized: better loop on every row VS all data load to RAM
                    def search_image(sheet, r, c):
                        for img in sheet._images:
                            if img.anchor._from.row == r and img.anchor._from.col == c:
                                return img._data()
                        return None
                    rdr = reader.excel.ExcelReader(file, False, False, False, True, False)
                    if not PILImage:
                        reader.excel.find_images = raw_find_images
                    rdr.read()
                    for sheetname in rdr.wb.sheetnames:
                        self.logi('💡SHEET NAME', sheetname)
                        ws = rdr.wb[sheetname]
                        self.logi('🎈IMAGES', len(ws._images))
                        list_rows = list(ws.rows)
                        prefix_cells = list_rows[0]
                        self.logi(prefix_cells)
                        row = 0
                        for cells in list_rows[1:]:
                            if msg_err and msg_err[-1] != '\n':
                                msg_err += '\n'
                            row += 1
                            p_id, p_name, p_image = 0, '', None
                            col = 0
                            try:
                                for cell in list(cells):
                                    match col:
                                        case 0:
                                            p_id = cell.value
                                        case 1:
                                            p_name = cell.value.strip() if cell.value else ''
                                        case 2:
                                            #🕵🔍 slowly, but memory optimized
                                            p_image = search_image(ws, row, col)
                                        case _:
                                            break
                                    col += 1
                            except Exception as e:
                                self.loge(e)
                                msg_err += f'ROW {row}: {e}'
                            else:
                                if p_name and p_image:
                                    product = Product.objects.filter(name=p_name).order_by('id').first()
                                    if product and not product.thumbnail and p_image:
                                        idata = base64.b64encode(p_image).decode()
                                        product.thumbnail = f'data:image/{idata[3:5].lower()};base64,{idata}'
                                        updated_products.append(product)
                    if updated_products:
                        try:
                            Product.objects.bulk_update(updated_products, ['thumbnail'])
                        except Exception as e:
                            self.loge(e)
                            updated_products = []
                self.message_user(request, f'🆗 {file.name} ✏️ FILE SIZE={file.size} ✏️ UPDATED={len(updated_products)}; {msg_err}', messages.SUCCESS)
                return HttpResponseRedirect(request.get_full_path())
        if not form:
            form = UploadFileForm(initial={'_selected_action': request.POST.getlist(admin.helpers.ACTION_CHECKBOX_NAME)})
        m = queryset.model._meta
        context = {}
        context['items'] = []
        context['form'] = form
        context['title'] = _('File')
        context['current_action'] = sys._getframe().f_code.co_name
        context['subtitle'] = 'admin_select_file_form'
        context['site_title'] = queryset.model._meta.verbose_name
        context['is_popup'] = True
        context['is_nav_sidebar_enabled'] = True
        context['site_header'] = _('Admin panel')
        context['has_permission'] = True
        context['site_url'] = reverse('admin:{}_{}_changelist'.format(m.app_label, m.model_name))
        context['available_apps'] = (m.app_label,)
        context['app_label'] = m.app_label
        return render(request, 'admin_select_file_form.html', context)
    thumbnails_from_xls.short_description = f'📄⚔✏️{_("load thumbnails from XLS file")}📄'

    def from_xls_with_check(self, request, queryset):
        from barcode import EAN13
        from transliterate import slugify
        from openpyxl import load_workbook
        form = None
        if 'apply' in request.POST:
            self.logi('💡', request.FILES)
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                msg_err = ''
                count_created = 0
                file = form.cleaned_data['file']
                if file:
                    units, groups, doc_income = {}, {}, None
                    timestamp = int(time.time())
                    wb = load_workbook(file)
                    for sheetname in wb.sheetnames:
                        self.logi('💡SHEET NAME', sheetname)
                        ws = wb[sheetname]
                        list_rows = list(ws.rows)
                        prefix_cells = list_rows[0]
                        self.logi(prefix_cells)
                        row_index = 0
                        for row in list_rows[1:]:
                            if msg_err and msg_err[-1] != '\n':
                                msg_err += '\n'
                            row_index += 1
                            p_group, p_code, p_name, p_article, p_unit, p_price, p_cost, p_barcode, p_count = '', 0, '', '', '', 0, 0, '', 0
                            i = 0
                            try:
                                for it in list(row):
                                    match i:
                                        case 0:
                                            p_group = it.value
                                        case 1:
                                            p_code = it.value
                                        case 2:
                                            p_name = it.value
                                        case 3:
                                            p_article = it.value
                                        case 4:
                                            p_unit = it.value
                                        case 5:
                                            p_price = it.value
                                            if isinstance(p_price, str):
                                                try:
                                                    p_price = Decimal(p_price)
                                                except Exception as e:
                                                    self.loge(e)
                                        case 6:
                                            p_cost = it.value
                                            if isinstance(p_cost, str):
                                                try:
                                                    p_cost = Decimal(p_cost)
                                                except Exception as e:
                                                    self.loge(e)
                                        case 7:
                                            p_barcode = it.value
                                        case 8:
                                            p_count = it.value
                                        case _:
                                            break
                                    i += 1
                            except Exception as e:
                                self.loge(e)
                                msg_err += f'ROW {row_index}: {e}'
                            else:
                                if not p_article and hasattr(settings, 'NEW_ARTICLE_GENERATOR'):
                                    p_article = settings.NEW_ARTICLE_GENERATOR(Product, f'{round(time.time())}')
                                if Product.objects.filter(Q(name=p_name) | Q(article=p_article)).exists():
                                    self.logi('PRODUCT', p_article, p_name, 'EXISTS')
                                else:
                                    product_kwargs = {'article':p_article if p_article else f'{timestamp+row_index}', 'name':p_name, 'cost':p_cost if isinstance(p_cost, (int, float, Decimal)) else 0, 'price':p_price if isinstance(p_price, (int, float, Decimal)) else 0}
                                    if p_unit:
                                        if p_unit not in units:
                                            condition_unit = Q(label=p_unit) | Q(label__icontains=p_unit)
                                            condition_unit |= Q(name=p_unit) | Q(name__icontains=p_unit)
                                            unit = Unit.objects.filter(condition_unit).first()
                                            if not unit:
                                                unit = Unit(label=p_unit, name=p_unit)
                                                try:
                                                    unit.save()
                                                except Exception as e:
                                                    self.loge(e)
                                                    unit = None
                                            if unit:
                                                units[p_unit] = unit
                                        if p_unit in units:
                                            product_kwargs['unit'] = units[p_unit]
                                    if p_group:
                                        if p_group not in groups:
                                            groups[p_group], created_group = ProductGroup.objects.get_or_create(name__icontains=p_group, defaults={'name':p_group})
                                        product_kwargs['group'] = groups[p_group]
                                    p = Product(**product_kwargs)
                                    try:
                                        p.save()
                                    except Exception as e:
                                        self.loge(e)
                                        msg_err += f'; {e}'
                                        continue
                                    count_created += 1
                                    time.sleep(.01)#FOR STRONG NEXT EAN13 GENERATION
                                    barcode = f'{p_barcode}' if p_barcode is not None else EAN13(f'{round(time.time()*1000)}').ean
                                    if barcode:
                                        b, created = BarCode.objects.get_or_create(id=barcode)
                                        if b:
                                            p.barcodes.add(b)
                                    if p_count:
                                        if not doc_income:
                                            t, created = DocType.objects.get_or_create(alias='balance', defaults={'alias':'balance', 'name':'Balance'})
                                            doc_income = get_model('core.Doc')(type=t, author=request.user)
                                            try:
                                                doc_income.save()
                                            except Exception as e:
                                                self.loge(e)
                                                msg_err += f'; {e}'
                                        if doc_income:
                                            r = get_model('core.Record')(count=p_count, cost=p.cost, price=p.price, doc=doc_income, product=p)
                                            try:
                                                r.save()
                                            except Exception as e:
                                                self.loge(e)
                                                msg_err += f'; {e}'
                                            else:
                                                try:
                                                    get_model('core.Register')(rec=r).save()
                                                except Exception as e:
                                                    self.loge(e)
                                                    msg_err += f'; {e}'
                self.message_user(request, f'🆗 {file.name} ✏️ FILE SIZE={file.size} ✏️ CREATED={count_created}; {msg_err}', messages.SUCCESS)
                return HttpResponseRedirect(request.get_full_path())
        if not form:
            form = UploadFileForm(initial={'_selected_action': request.POST.getlist(admin.helpers.ACTION_CHECKBOX_NAME)})
        m = queryset.model._meta
        context = {}
        context['items'] = []
        context['form'] = form
        context['title'] = _('File')
        context['current_action'] = sys._getframe().f_code.co_name
        context['subtitle'] = 'admin_select_file_form'
        context['site_title'] = queryset.model._meta.verbose_name
        context['is_popup'] = True
        context['is_nav_sidebar_enabled'] = True
        context['site_header'] = _('Admin panel')
        context['has_permission'] = True
        context['site_url'] = reverse('admin:{}_{}_changelist'.format(m.app_label, m.model_name))
        context['available_apps'] = (m.app_label,)
        context['app_label'] = m.app_label
        return render(request, 'admin_select_file_form.html', context)
    from_xls_with_check.short_description = f'📍⚔📉{_("load from XLS file with check (slow)")}📈'

    def from_xls(self, request, queryset):
        from barcode import EAN13
        from transliterate import slugify
        from openpyxl import load_workbook
        form = None
        if 'apply' in request.POST:
            self.logi('💡', request.FILES)
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                msg_err = ''
                count_created = 0
                file = form.cleaned_data['file']
                if file:
                    units, groups, ext_attrs, products = {}, {}, {}, []
                    timestamp = int(time.time())
                    wb = load_workbook(file)
                    for sheetname in wb.sheetnames:
                        self.logi('💡SHEET NAME', sheetname)
                        ws = wb[sheetname]
                        list_rows = list(ws.rows)
                        prefix_cells = list_rows[0]
                        self.logi(prefix_cells)
                        row_index = 0
                        for row in list_rows[1:]:
                            row_index += 1
                            p_group, p_code, p_name, p_article, p_unit, p_price, p_cost, p_barcode, p_count = '', 0, '', '', '', 0, 0, '', 0
                            i = 0
                            try:
                                for it in list(row):
                                    match i:
                                        case 0:
                                            p_group = it.value
                                        case 1:
                                            p_code = it.value
                                        case 2:
                                            p_name = it.value
                                        case 3:
                                            p_article = it.value
                                        case 4:
                                            p_unit = it.value
                                        case 5:
                                            p_price = it.value
                                        case 6:
                                            p_cost = it.value
                                        case 7:
                                            p_barcode = it.value
                                        case 8:
                                            p_count = it.value
                                        case _:
                                            break
                                    i += 1
                            except Exception as e:
                                self.loge(e)
                                msg_err += f'ROW {row_index}: {e}'
                            else:
                                if Product.objects.filter(Q(name=p_name) | Q(article=p_article)).exists():
                                    self.logi('PRODUCT', p_article, p_name, 'EXISTS')
                                else:
                                    product_kwargs = {'article':p_article if p_article else f'{timestamp+row_index}', 'name':p_name, 'cost':p_cost if isinstance(p_cost, (int, float)) else 0, 'price':p_price if isinstance(p_price, (int, float)) else 0}
                                    if p_unit:
                                        if p_unit not in units:
                                            condition_unit = Q(label=p_unit) | Q(label__icontains=p_unit)
                                            condition_unit |= Q(name=p_unit) | Q(name__icontains=p_unit)
                                            unit = Unit.objects.filter(condition_unit).first()
                                            if not unit:
                                                unit = Unit(label=p_unit, name=p_unit)
                                                try:
                                                    unit.save()
                                                except Exception as e:
                                                    self.loge(e)
                                                    unit = None
                                            if unit:
                                                units[p_unit] = unit
                                        if p_unit not in units:
                                            product_kwargs['unit'] = units[p_unit]
                                    if p_group:
                                        if p_group not in groups:
                                            groups[p_group], created_group = ProductGroup.objects.get_or_create(name__icontains=p_group, defaults={'name':p_group})
                                        product_kwargs['group'] = groups[p_group]
                                    p = Product(**product_kwargs)
                                    products.append(p)
                                    ext_attrs[p.article] = {'barcode':f'{p_barcode if p_barcode else EAN13(f'{round(time.time()*1000)}').ean}', 'count':p_count}
                                    time.sleep(.01)#FOR STRONG NEXT EAN13 GENERATION
                    if products:
                        try:
                            objs = Product.objects.bulk_create(products)
                        except Exception as e:
                            self.loge(e)
                            msg_err = f'{e}'
                        else:
                            count_created = len(objs)
                            doc_income, income_mybe_saved = None, True
                            if count_created:
                                for o in objs:
                                    barcode = ext_attrs[o.article]['barcode']
                                    if barcode:
                                        b, created = BarCode.objects.get_or_create(id=barcode)
                                        if b:
                                            o.barcodes.add(b)
                                    p_count = ext_attrs[o.article]['count']
                                    if p_count and income_mybe_saved:
                                        if not doc_income:
                                            t, created = DocType.objects.get_or_create(alias='balance', defaults={'alias':'balance', 'name':'Balance'})
                                            doc_income = get_model('core.Doc')(type=t, author=request.user)
                                            try:
                                                doc_income.save()
                                            except Exception as e:
                                                self.loge(e)
                                                income_mybe_saved = False
                                                doc_income = None
                                        if doc_income:
                                            r = get_model('core.Record')(count=p_count, cost=o.cost, price=o.price, doc=doc_income, product=o)
                                            try:
                                                r.save()
                                            except Exception as e:
                                                self.loge(e)
                                            else:
                                                try:
                                                    get_model('core.Register')(rec=r).save()
                                                except Exception as e:
                                                    self.loge(e)
                self.message_user(request, f'🆗 {file.name} ✏️ FILE SIZE={file.size} ✏️ CREATED={count_created}; {msg_err}', messages.SUCCESS)
                return HttpResponseRedirect(request.get_full_path())
        if not form:
            form = UploadFileForm(initial={'_selected_action': request.POST.getlist(admin.helpers.ACTION_CHECKBOX_NAME)})
        m = queryset.model._meta
        context = {}
        context['items'] = []
        context['form'] = form
        context['title'] = _('File')
        context['current_action'] = sys._getframe().f_code.co_name
        context['subtitle'] = 'admin_select_file_form'
        context['site_title'] = queryset.model._meta.verbose_name
        context['is_popup'] = True
        context['is_nav_sidebar_enabled'] = True
        context['site_header'] = _('Admin panel')
        context['has_permission'] = True
        context['site_url'] = reverse('admin:{}_{}_changelist'.format(m.app_label, m.model_name))
        context['available_apps'] = (m.app_label,)
        context['app_label'] = m.app_label
        return render(request, 'admin_select_file_form.html', context)
    from_xls.short_description = f'⚔{_("load from XLS file")}'

    def to_xls(self, request, queryset):
        queryset = queryset.annotate(unit_label=F('unit__label'), barcode_first=F('barcodes__id'), qrcode_first=F('qrcodes__id'))
        columns = {
            'group':{'width':20},
            'id':{'width':10},
            'name':{'width':20},
            'article':{'width':20},
            'unit_label':{'title':'unit'},
            'price':{},
            'cost':{},
            'barcode_first':{'width':20, 'title':'barcode'},
            'qrcode_first':{'width':20, 'title':'qrcode'}
        }
        output = self.queryset_to_xls(queryset, columns)
        if output:
            fn = '{}.xlsx'.format(django_timezone.now().strftime('%Y%m%d%H%M%S'))
            self.message_user(request, f'🆗 {_("Finished")} ✏️({fn})', messages.SUCCESS)
            return FileResponse(output, as_attachment=True, filename=fn)
        self.message_user(request, _('please select items'), messages.ERROR)
    to_xls.short_description = f'⚔{_("export to XLS file")}↘'

    def price_to_xls(self, request, queryset):
        last_register = get_model('core.Register').objects.filter(rec__product_id=OuterRef('pk')).order_by('-rec__doc__registered_at')[:1]
        queryset = queryset.annotate(last_price=Subquery(last_register.values('rec__price')))
        output = self.queryset_to_xls(queryset, {'article':{'width':30}, 'name':{'width':50}, 'last_price':{'width':20, 'title':'price'}})
        if output:
            fn = '{}.xlsx'.format(django_timezone.now().strftime('%Y%m%d%H%M%S'))
            self.message_user(request, f'🆗 {_("Finished")} ✏️({fn})', messages.SUCCESS)
            return FileResponse(output, as_attachment=True, filename=fn)
        self.message_user(request, _('please select items'), messages.ERROR)
    price_to_xls.short_description = f'⚔{_("unload price to XLS file")}↘'

    def thumbnails_to_xls(self, request, queryset):
        columns = {
            'id':{'width':10},
            'name':{'width':20},
            'thumbnail':{'width':64}
        }
        output = self.queryset_to_xls(queryset, columns)
        if output:
            fn = '{}.xlsx'.format(django_timezone.now().strftime('%Y%m%d%H%M%S'))
            self.message_user(request, f'🆗 {_("Finished")} ✏️({fn})', messages.SUCCESS)
            return FileResponse(output, as_attachment=True, filename=fn)
        self.message_user(request, _('please select items'), messages.ERROR)
    thumbnails_to_xls.short_description = f'📄⚔{_("export thumbnails to XLS file")}↘📄'

    def barcode_to_svg(self, request, queryset):
        from textwrap import wrap
        from barcode import EAN13
        from barcode.writer import SVGWriter
        def eval_dim(val, n, m='-'):
            v = re.findall(r'\d+\.\d+', val)[0]
            suffix = val.replace(v, '')
            return f'{eval(f'{v}{m}{n}')}{suffix}'
        svgs = ''
        svg_width = '30mm'
        svg_height = '20mm'
        font_size = settings.ADMIN_EAN13_RENDER_OPTIONS.get('font_size', 8)
        font_style_ext = settings.ADMIN_EAN13_RENDER_OPTIONS.get('font_style_ext', '')
        text_wrapped_symbols = settings.ADMIN_EAN13_RENDER_OPTIONS.get('text_wrapped_symbols', 15)
        show_name = settings.ADMIN_EAN13_RENDER_OPTIONS.get('show_name', False)
        name_at_top = settings.ADMIN_EAN13_RENDER_OPTIONS.get('name_at_top', False)
        name_at_top_h = settings.ADMIN_EAN13_RENDER_OPTIONS.get('name_at_top_h', font_size)
        name_at_top_x = settings.ADMIN_EAN13_RENDER_OPTIONS.get('name_at_top_x', 'center')
        css_media_orientation = settings.ADMIN_EAN13_RENDER_OPTIONS.get('css_media_orientation', '')
        css_media_page_size_ext = settings.ADMIN_EAN13_RENDER_OPTIONS.get('css_media_page_size_ext', 'landscape;page-orientation:rotate-right')
        css_media_ext = settings.ADMIN_EAN13_RENDER_OPTIONS.get('css_media_ext', '')
        print_button = settings.ADMIN_EAN13_RENDER_OPTIONS.get('print_button', '')
        print_script = settings.ADMIN_EAN13_RENDER_OPTIONS.get('print_script', '')
        for it in queryset:
            bcode_value = it.barcodes.first()
            if bcode_value:
                name_wrapped = '\n'.join(wrap(it.name, text_wrapped_symbols))
                svgwriter = SVGWriter()
                ean = EAN13(bcode_value.id, writer=svgwriter)
                if show_name:
                    svg = ean.render(settings.ADMIN_EAN13_RENDER_OPTIONS, name_wrapped).decode('UTF-8').replace('\n', '')
                else:
                    svg = ean.render(settings.ADMIN_EAN13_RENDER_OPTIONS).decode('UTF-8').replace('\n', '')
                if settings.DEBUG:
                    self.logi(bcode_value.id, ean.to_ascii())
                if name_at_top and not show_name:
                    svg_width = svgwriter._root.getAttribute('width')
                    svg_height = f'{float(svgwriter._root.getAttribute('height').replace('mm', '')) + name_at_top_h}mm'
                    if name_at_top_x == 'center':
                        texts = svgwriter._document.getElementsByTagName('text')
                        if texts:
                            name_at_top_x = texts[0].getAttribute('x')
                        else:
                            name_at_top_x = '0mm'
                    name_wrapped = f'</tspan><tspan x="{name_at_top_x}" dy="{name_at_top_h}pt">'.join(wrap(it.name, text_wrapped_symbols)).join([f'<tspan x="{name_at_top_x}" dy="{name_at_top_h}pt">','</tspan>'])
                    svg_top = f'<svg id="top" xmlns="http://www.w3.org/2000/svg" width="{eval_dim(svg_width,.5)}" height="{eval_dim(svg_height,1.911)}"><g id="top-g"><text id="top-text" x="{name_at_top_x}" style="font-size:{font_size}pt;text-anchor:middle;{font_style_ext}">{name_wrapped}</text></g>'
                    svgs += f'<p class="page-pad">{svg_top}{svg.replace('<svg', f'<svg y="{name_at_top_h}mm"')}</svg></p>'
                else:
                    svgs += f'<p class="page-pad">{svg}</p>'
        svgs = f'<div id="section-to-print"><style>@media {css_media_orientation} print{{html body{{visibility:hidden;height:auto;margin:0;padding:0;}} .content{{position:absolute;top:0;}} .messagelist{{margin:0;padding:0;}} #section-to-print{{text-align:center;background-color:white;width:0;display:flex;flex-direction:column;visibility:visible;position:absolute;left:0;top:0;}}}} @page{{size: {svg_width} {svg_height} {css_media_page_size_ext};margin:0;}} .page-pad{{break-after:page;margin:0;padding:0;}} .page-pad:last-of-type{{break-after:avoid!important;}}{css_media_ext}</style>{print_button}{print_script}' + re.sub('(<!--.*?-->)', '', svgs, flags=re.DOTALL) + '</div>'
        self.message_user(request, mark_safe(svgs), messages.SUCCESS)
    barcode_to_svg.short_description = f'🖶{_("print barcode as SVG")}🖼'

    def fix_barcodes(self, request, queryset):
        from .models import ean13
        msg = ''
        delete_bcodes = []
        fixcount = 0
        for it in queryset:
            for bcode in it.barcodes.all():
                e13 = ean13(bcode.id)
                if e13 != bcode.id:
                    if settings.DEBUG:
                        self.logw('INVALID', bcode.id)
                    it.barcodes.remove(bcode)
                    delete_bcodes.append(bcode.id)
                    if BarCode.objects.filter(id=e13).exists():
                        e13 = ean13(f'{round(time.time()*1000)}')
                    bcode = BarCode(e13)
                    try:
                       bcode.save()
                    except Exception as e:
                       self.loge(e)
                    else:
                        it.barcodes.add(bcode)
                        if settings.DEBUG:
                            self.logw('FIXED TO', bcode.id)
                        fixcount += 1
        if fixcount:
            msg += f'FIXED {fixcount}'
        if delete_bcodes:
            try:
                dresult = BarCode.objects.filter(id__in=delete_bcodes).delete()
            except Exception as e:
                self.loge(e)
            else:
                m = f'{_("delete barcodes")} {dresult}'
                if settings.DEBUG:
                    self.logi(m)
                msg += f'; {m}'
        self.message_user(request, msg, messages.SUCCESS)
    fix_barcodes.short_description = f'Ⅲ🔨 {_("fix barcodes")} 🔧'

    def copy_unit(self, request, queryset):
        changed = 0
        if queryset.count() > 1:
            it_first = queryset.first()
            changed = queryset.exclude(unit_id=it_first.unit_id).update(unit_id=it_first.unit_id)
        self.message_user(request, f'{_("copied")} {changed}', messages.SUCCESS)
    copy_unit.short_description = f'🆔 {_("copy unit")} 🆗'

    def copy_cost(self, request, queryset):
        changed = 0
        if queryset.count() > 1:
            it_first = queryset.first()
            changed = queryset.exclude(cost=it_first.cost).update(cost=it_first.cost)
        self.message_user(request, f'{_("copied")} {changed}', messages.SUCCESS)
    copy_cost.short_description = f'💲 {_("copy cost")} ₿'

    def copy_price(self, request, queryset):
        changed = 0
        if queryset.count() > 1:
            it_first = queryset.first()
            changed = queryset.exclude(price=it_first.price).update(price=it_first.price)
        self.message_user(request, f'{_("copied")} {changed}', messages.SUCCESS)
    copy_price.short_description = f'💰 {_("copy price")} ₿'

    def copy_cost_price(self, request, queryset):
        changed = 0
        if queryset.count() > 1:
            it_first = queryset.first()
            changed = queryset.exclude(cost=it_first.cost, price=it_first.price).update(cost=it_first.cost, price=it_first.price)
        self.message_user(request, f'{_("copied")} {changed}', messages.SUCCESS)
    copy_cost_price.short_description = f'🪙 {_("copy cost and price")} ₿💲💰'

    def order_from_selected_items(self, request, queryset):
        errs = ''
        if queryset.count() > 1:
            get_model('core.Doc')()
            als = 'order'
            doc_type, created = DocType.objects.get_or_create(alias=als, defaults={'alias':als, 'name':als})
            id_owner = request.user.default_company.id if request.user.default_company else 1
            doc = get_model('core.Doc')(type=doc_type, owner_id=id_owner, author=request.user)
            try:
                doc.save()
            except Exception as e:
                self.loge(e)
                errs += f'{e}'
            else:
                recs = [get_model('core.Record')(count=1, cost=p.cost, price=p.price, doc=doc, currency=p.currency, product=p) for p in queryset]
                try:
                    obj_recs = get_model('core.Record').objects.bulk_create(recs)
                except Exception as e:
                    self.loge(e, doc, recs)
                return HttpResponseRedirect(f'{settings.ADMIN_PATH_PREFIX}/core/doc/{doc.id}/change/')
        self.message_user(request, errs, messages.ERROR)
    order_from_selected_items.short_description = f'🗂 {_("create order from selected products")}'

    def qr_to_svg(self, request, queryset):
        template, created = get_model('refs.PrintTemplates').objects.get_or_create(alias='refs.qrcode', defaults=settings.DEAFAULT_QR_PRINT_TEMPLATE)
        if not template:
            self.message_user(request, _('please add qrcode printer template first'), messages.ERROR)
            return
        import qrcode
        import qrcode.image.svg
        svgs = ''
        for it in queryset:
            code_value = it.qrcodes.first()
            if not code_value:
                code_value = it.barcodes.first()
            if code_value:
                svg = qrcode.make(code_value.id, image_factory=qrcode.image.svg.SvgPathImage, border=template.extinfo.get('qr_border', 0), box_size=template.extinfo.get('qr_box_size', 10), version=template.extinfo.get('qr_version', 1))
                svgs += Template(template.content).render(Context({'svg':svg.to_string().decode('utf-8')}))
        if 'script' in template.extinfo:
            svgs += template.extinfo['script']
        svgs = f'<div id="section-to-print"><style>{template.extinfo["css_media_style"]}</style>' + re.sub('(<!--.*?-->)', '', svgs, flags=re.DOTALL) + f'</div>'
        self.message_user(request, mark_safe(svgs), messages.SUCCESS)
    qr_to_svg.short_description = f'🖶{_("print QR as SVG")}㊙️'

    def copy_name_to_ext_label(self, request, queryset):
        msg = '-'
        updated_items = []
        for it in queryset:
            label = it.extinfo.get('label', None)
            if not label:
                it.extinfo['label'] = it.name
                updated_items.append(it)
        if updated_items:
            msg = _('updated')
            try:
                res = Product.objects.bulk_update(updated_items, ['extinfo'])
            except Exception as e:
                self.loge(e)
                msg += f' {e}'
            else:
                msg += f' {res}'
        self.message_user(request, mark_safe(msg), messages.SUCCESS)
    copy_name_to_ext_label.short_description = f'🏷️{_("copy name to extra label")}🏷️'

admin.site.register(Product, ProductAdmin)


class CustomerAdmin(CustomModelAdmin):
    list_display = ('id', 'name', 'extinfo')
    list_display_links = ('id', 'name')
    search_fields = ('id', 'name', 'extinfo')
admin.site.register(Customer, CustomerAdmin)


class ProductImageAdmin(CustomModelAdmin):
    list_display = ('id', 'created_at', 'file', 'extinfo')
    list_display_links = ('id', 'created_at')
    search_fields = ('id', 'created_at', 'file', 'extinfo')
admin.site.register(ProductImage, ProductImageAdmin)
