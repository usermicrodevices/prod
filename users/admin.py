import sys, logging
from io import BytesIO
from datetime import datetime, timedelta

from django.contrib import admin, messages

from django import forms
from django.urls import reverse
from django.conf import settings
from django.shortcuts import render
from django.apps import apps as django_apps
from django.db.models import Q, Value, CharField
from django.utils import timezone as django_timezone
from django.utils.safestring import mark_safe
from django.utils.html import format_html, format_html_join
from django.utils.translation import gettext as _
from django.http import StreamingHttpResponse, FileResponse

from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.contrib.sessions.models import Session
#from django.contrib.sessions.backends.db import SessionStore

from .models import get_users_by_owner, Role, RoleModel, RoleField, User

admin.site.subtitle = _('Users')

def get_model(app_model):
    app_name, model_name = app_model.split('.')
    return django_apps.get_app_config(app_name).get_model(model_name)

class UploadFileForm(forms.Form):
    _selected_action = forms.CharField(widget=forms.MultipleHiddenInput)
    file = forms.FileField(widget=forms.ClearableFileInput(attrs={'allow_multiple_selected': True}))


class SessionAdmin(admin.ModelAdmin):
    _session_store_ = Session.get_session_store_class()()
    list_display = ('expire_date', 'session_key', 'get_session_data')
    search_fields = ('expire_date', 'session_key', 'session_data')

    def get_session_data(self, obj):
        return self._session_store_.decode(obj.session_data) if obj.session_data else ''
    get_session_data.admin_order_field = 'session_data'
admin.site.register(Session, SessionAdmin)


class LogEntryAdmin(admin.ModelAdmin):
    list_display = ('id', 'action_time', 'user', 'content_type', 'object_repr', 'action_flag', 'change_message')
    search_fields = ('id', 'action_time', 'user__username', 'user__first_name', 'user__last_name', 'user__email', 'object_repr', 'change_message')
admin.site.register(admin.models.LogEntry, LogEntryAdmin)


class ContentTypeAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'model', 'app_label')
    list_display_links = ['name']
    search_fields = ['id', 'model', 'app_label']
    list_filter = ['app_label']
admin.site.register(ContentType, ContentTypeAdmin)


class RoleAdmin(admin.ModelAdmin):
    list_display = ('id', 'weight', 'value', 'description', 'get_group')
    list_display_links = ['id', 'value']
    search_fields = ['value', 'description', 'group__name']
    list_select_related = ['group']
    list_filter = ['group']

    def get_group(self, obj):
        grp = obj.group
        if not grp:
            return ''
        model_meta = grp._meta
        info = (model_meta.app_label, model_meta.model_name)
        admin_url = reverse('admin:%s_%s_change' % info, args=[obj.group.pk])
        return format_html('<font size="+1"><a href="{}" target="_blank">{}</a></font>', admin_url, obj.group)
    get_group.short_description = _('group')
    get_group.admin_order_field = 'group'

admin.site.register(Role, RoleAdmin)


class RoleModelAdmin(admin.ModelAdmin):
    list_display = ('id', 'app', 'model')
    list_display_links = ['id', 'app', 'model']
    search_fields = ['app', 'model']
admin.site.register(RoleModel, RoleModelAdmin)


class RoleFieldAdmin(admin.ModelAdmin):
    list_display = ('id', 'value', 'role', 'role_model', 'read', 'write')
    list_display_links = ['value']
    search_fields = ['value', 'role__value', 'role_model__value']
    list_select_related = ['role', 'role_model']
    list_filter = ['role', 'role_model']
    actions = ('set_read', 'reset_read', 'set_write', 'reset_write')

    def set_read(self, request, queryset):
        queryset.update(read=True)
    set_read.short_description = f'☑{_("set role field for read")}√'

    def reset_read(self, request, queryset):
        queryset.update(read=False)
    reset_read.short_description = f'☒{_("make role field unreadable")}☐'

    def set_write(self, request, queryset):
        queryset.update(write=True)
    set_write.short_description = f'☑{_("set role field for write")}√'

    def reset_write(self, request, queryset):
        queryset.update(write=False)
    reset_write.short_description = f'☒{_("make role field unwritable")}☐'

admin.site.register(RoleField, RoleFieldAdmin)


class PermissionAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'content_type', 'codename')
    list_display_links = ['name']
    search_fields = ['id', 'name', 'codename']
admin.site.register(Permission, PermissionAdmin)


class UserAdmin(BaseUserAdmin):
    list_display = ('id', 'get_avatar', 'staff', 'get_username', 'email', 'first_name', 'last_name', 'get_role', 'last_login', 'get_groups', 'get_companies', 'get_sale_points', 'default_company')
    search_fields = ('username', 'email', 'first_name', 'last_name', 'last_login', 'role__value', 'role__description', 'companies__name', 'sale_points__name')
    list_select_related = ('role',)
    list_filter = ('is_staff', 'is_superuser', 'is_active', 'role')
    fieldsets = (
        *BaseUserAdmin.fieldsets,
        (
            'Extended',
            {
                'fields': (
                    'role',
                    'default_company',
                    'companies',
                    'sale_points',
                    'avatar'
                ),
            },
        ),
    )
    actions = ('load_from_xls', 'set_all_push_notifications', 'selected_to_xls', 'filtered_by_selected_to_xls', 'add_notifies_by_last_from_same_notify_group_and_company')

    def logw(self, *args):
        msg += f'⚠️{self.__class__.__name__}.{sys._getframe().f_back.f_code.co_name}'
        for arg in args:
            msg += f'::{arg}'
        logging.warning(msg)

    def logi(self, *args):
        msg += f'💡{self.__class__.__name__}.{sys._getframe().f_back.f_code.co_name}'
        for arg in args:
            msg += f'::{arg}'
        logging.info(msg)

    def loge(self, err, *args):
        msg = f'🆘{self.__class__.__name__}.{err.__traceback__.tb_frame.f_code.co_name}'
        for arg in args:
            msg += f'::{arg}'
        msg += f'::{err}::LINE={err.__traceback__.tb_lineno}'
        logging.error(msg)

    def worksheet_cell_write(self, worksheet, row, col, value, type_value = None, fmt = None):
        func_write = worksheet.write
        if type_value == 'as_number':
            func_write = worksheet.write_number
        elif type_value == 'as_datetime':
            func_write = worksheet.write_datetime
        try:
            if fmt:
                func_write(row, col, value, fmt)
            else:
                func_write(row, col, value)
        except Exception as e:
            try:
                if fmt:
                    func_write(row, col, repr(value), fmt)
                else:
                    func_write(row, col, repr(value))
            except Exception as e:
                self.loge(e, row, col)
        return col + 1

    def worksheet_row_write(self, worksheet, row, values):
        col = 0
        for item in values:
            col = self.worksheet_cell_write(worksheet, row, col, item)
        return row + 1

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        user = request.user
        if user.is_superuser:
            return qs
        return qs.filter(Q(companies__in=user.companies.all()) | Q(sale_points__in=user.sale_points.all())).exclude(role__weight__lt=user.role.weight).distinct()

    def staff(self, obj):
        return '✅' if obj.is_staff else '❌'
    staff.short_description = _('Staff')
    staff.admin_order_field = 'is_staff'

    def get_username(self, obj):
        model_meta = obj._meta
        info = (model_meta.app_label, model_meta.model_name)
        admin_url = reverse('admin:%s_%s_change' % info, args=(obj.pk,))
        return format_html('<font size="+1"><a href="{}" target="_blank">{}</a></font>', admin_url, obj.username)
    get_username.short_description = _('Login')
    get_username.admin_order_field = 'username'

    def get_role(self, obj):
        if obj.role:
            model_meta = obj.role._meta
            info = (model_meta.app_label, model_meta.model_name)
            admin_url = reverse('admin:%s_%s_change' % info, args=(obj.role_id,))
            return format_html('<p><font color="gray" face="Verdana, Geneva, sans-serif"><a href="{}" target="_blank">{}</a></font></p>', admin_url, obj.role.value)
        else:
            return ''
    get_role.short_description = _('Role')
    get_role.admin_order_field = 'role'

    def get_groups(self, obj):
        return format_html_join('\n', '<p><font color="gray" face="Verdana, Geneva, sans-serif"><a href="{}/auth/group/{}/change/" target="_blank">{}</a></font></p>', obj.groups.annotate(admin_path_prefix=Value(settings.ADMIN_PATH_PREFIX, CharField())).values_list('admin_path_prefix', 'id', 'name'))
    get_groups.short_description = _('Groups')
    get_groups.admin_order_field = 'groups'

    def get_companies(self, obj):
        return format_html_join('\n', '<p><font color="green" face="Verdana, Geneva, sans-serif"><a href="{}/core/company/{}/change/" target="_blank">{}</a></font></p>', obj.companies.annotate(admin_path_prefix=Value(settings.ADMIN_PATH_PREFIX, CharField())).values_list('admin_path_prefix', 'id', 'name'))
    get_companies.short_description = _('Companies')
    get_companies.admin_order_field = 'companies'

    def get_sale_points(self, obj):
        return format_html_join('\n', '<p><font color="green" face="Verdana, Geneva, sans-serif"><a href="{}/core/salepoint/{}/change/" target="_blank">{}</a></font></p>', obj.sale_points.annotate(admin_path_prefix=Value(settings.ADMIN_PATH_PREFIX, CharField())).values_list('admin_path_prefix', 'id', 'name'))
    get_sale_points.short_description = _('Sale Points')
    get_sale_points.admin_order_field = 'sale_points'

    def get_avatar(self, obj):
        if obj.avatar:
            return format_html('<img src="{}" width="32" height="32">', obj.avatar)
        else:
            default_content = '''data:image/svg+xml,<?xml version="1.0" encoding="UTF-8" standalone="no"?><!DOCTYPE svg PUBLIC "-//W3C//DTD SVG 1.1//EN" "http://www.w3.org/Graphics/SVG/1.1/DTD/svg11.dtd"><svg version="1.1" baseProfile="full" xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink" xmlns:ev="http://www.w3.org/2001/xml-events" width="100%" height="100%"><rect fill="white" x="0" y="0" width="100%" height="100%" /><rect fill="green" x="0" y="0" width="100%" height="100%" rx="1em"/></svg>'''
            return format_html('''<img src="{}" width="32" height="32">''', default_content)
    get_avatar.short_description = _('Avatar')

    def load_from_xls(self, request, queryset):
        import xlsxwriter
        from transliterate import slugify
        from openpyxl import load_workbook
        form = None
        if 'apply' in request.POST:
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = form.cleaned_data['file']
                if file:
                    if file.size > 1048576:#1MB
                        self.message_user(request, f'🚫ERROR BIG FILE SIZE = {file.size} BYTES', messages.ERROR)
                        return
                    ####################################
                    output = BytesIO()
                    fn = '{}.xlsx'.format(datetime.now().strftime('%Y%m%d%H%M%S'))
                    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
                    worksheet = workbook.add_worksheet()
                    row_id = 1
                    ####################################
                    companies_count = 0
                    sale_points_count = 0
                    devices_count = 0
                    wb = load_workbook(file)
                    for sheetname in wb.sheetnames:
                        self.logi(sheetname)
                        ws = wb[sheetname]
                        list_rows = list(ws.rows)
                        self.logi(list_rows[0][:10])
                        companies_dict = {}
                        current_company = ''
                        current_sp = ''
                        for row in list_rows[1:]:
                            try:
                                fl,zn,fio,rl,eml,phone,id_sps,company_ids,sale_point_names,company_id = list(row)[:10]
                            except Exception as e:
                                self.loge(e)
                            else:
                                try: fl = fl.value.strip()
                                except: fl = fl.value
                                try: zn = zn.value.strip()
                                except: zn = zn.value
                                try: fio = fio.value.strip()
                                except: fio = fio.value
                                try: rl = rl.value.strip()
                                except: rl = rl.value
                                try: eml = eml.value.strip()
                                except: eml = eml.value
                                try: phone = phone.value.strip()
                                except: phone = phone.value
                                try: id_sps = id_sps.value.strip()
                                except: id_sps = id_sps.value
                                if isinstance(id_sps, str) and ',' in id_sps:
                                    id_sps = [int(idsp) for idsp in id_sps.split(',')]
                                else:
                                    try:
                                        id_sps = [int(id_sps)]
                                    except:
                                        id_sps = []
                                if not isinstance(id_sps, list):
                                    id_sps = []
                                ####################
                                try: company_ids = company_ids.value.strip()
                                except: company_ids = company_ids.value
                                if isinstance(company_ids, str) and ',' in company_ids:
                                    company_ids = [int(idc) for idc in company_ids.split(',')]
                                else:
                                    try:
                                        company_ids = [int(company_ids)]
                                    except:
                                        company_ids = []
                                if not isinstance(company_ids, list):
                                    company_ids = []
                                ####################
                                if isinstance(sale_point_names, str) and ',' in sale_point_names:
                                    try: sale_point_names = sale_point_names.value.strip()
                                    except: sale_point_names = sale_point_names.value
                                    sale_point_names = sale_point_names.split(',')
                                else:
                                    sale_point_names = []
                                ####################
                                try:
                                    company_id = int(company_id.value.strip())
                                except:
                                    try:
                                        company_id = int(company_id.value)
                                    except:
                                        company_id = None
                                ####################
                                if not fl or len(fl) < 9 or not fio:
                                    continue
                                else:
                                    sfio = fio.split(' ')
                                    family = sfio[0]
                                    name, patronymic = '', ''
                                    if len(sfio) > 1:
                                        name = sfio[1]
                                    if len(sfio) > 2:
                                        patronymic = sfio[2]
                                    login = slugify(family)
                                    if User.objects.filter(username=login).count():
                                        login = f'{login}{User.objects.count()}'
                                    password = User.objects.make_random_password()
                                    user_args = [login, password, eml, family, name, patronymic]
                                    user = None
                                    try:
                                        user = User.objects.create_user(login, eml, password, first_name=name, last_name=family)
                                    except Exception as e:
                                        self.loge(e, user_args)
                                        login = f'{login}{User.objects.count()}'
                                        try:
                                            user = User.objects.create_user(login, eml, password, first_name=name, last_name=family)
                                        except Exception as e:
                                            self.loge(e, user_args)
                                    if not user:
                                        self.logw('USER-NOT-CREATED', user_args)
                                    else:
                                        try:
                                            role = Role.objects.get(value=rl)
                                        except Exception as e:
                                            self.loge(e)
                                            user.role_id = 2
                                        else:
                                            user.role = role
                                            user.groups.set([role.group])
                                        #user.patronymic_text = patronymic
                                        try:
                                            user.save()
                                        except Exception as e:
                                            self.loge(e)
                                        else:
                                            self.logi(user)
                                            for id_sp in id_sps:
                                                try:
                                                    sp = get_model('refs.SalePoint').objects.get(id=id_sp)
                                                except Exception as e:
                                                    self.loge(e)
                                                else:
                                                    try:
                                                        user.sale_points.add(sp)
                                                    except Exception as e:
                                                        self.loge(e)
                                            for idc in company_ids:
                                                try:
                                                    company = get_model('refs.Company').objects.get(id=idc)
                                                except Exception as e:
                                                    self.loge(e)
                                                else:
                                                    try:
                                                        user.companies.add(company)
                                                    except Exception as e:
                                                        self.loge(e)
                                            for spname in sale_point_names:
                                                search_kwargs = {'name__icontains':spname}
                                                if company_id:
                                                    search_kwargs['company_id'] = company_id
                                                self.loge(f'SEARCH-SALE-POINT {search_kwargs}')
                                                try:
                                                    sp = get_model('refs.SalePoint').objects.filter(**search_kwargs)[0]
                                                except Exception as e:
                                                    self.loge(e)
                                                else:
                                                    try:
                                                        user.sale_points.add(sp)
                                                    except Exception as e:
                                                        self.loge(e)
                                        values = [login, password, eml, family, name, patronymic, rl, zn]
                                        row_id = self.worksheet_row_write(worksheet, row_id, values)
                    ####################################
                    workbook.close()
                    output.seek(0)
                    ####################################
                    self.message_user(request, fn, messages.SUCCESS)
                    response = FileResponse(output, as_attachment=True, filename=fn)
                    return response
                else:
                    self.message_user(request, f'🚫ERROR LOAD FILE', messages.ERROR)
                    return
            else:
                self.message_user(request, f'🚫ERROR WITH FORM SELECT FILE', messages.ERROR)
                return
        if not form:
            form = UploadFileForm(initial={'_selected_action': request.POST.getlist(admin.helpers.ACTION_CHECKBOX_NAME)})
        m = queryset.model._meta
        context = {}
        context['items'] = queryset
        context['form'] = form
        context['title'] = _('File')
        context['current_action'] = sys._getframe().f_code.co_name
        context['subtitle'] = 'admin_select_file_form'
        context['site_title'] = _('Users')
        context['is_popup'] = True
        context['is_nav_sidebar_enabled'] = True
        context['site_header'] = _('Admin panel')
        context['has_permission'] = True
        context['site_url'] = reverse('admin:{}_{}_changelist'.format(m.app_label, m.model_name))
        context['available_apps'] = (m.app_label,)
        context['app_label'] = m.app_label
        return render(request, 'admin_select_file_form.html', context)
    load_from_xls.short_description = f'⚔️{_("load from XLSX file")}'

    def set_all_push_notifications(self, request, queryset):
        changed = 0
        try:
            push_type = get_model('refs.NotificationType').objects.get(value='push')
        except Exception as e:
            self.loge(e)
            self.message_user(request, f'{e}', messages.ERROR)
            return
        sources = get_model('refs.NotificationSource').objects.all()
        self.logi(f'Notification Sources = {sources.count()}')
        #self.logi(f'{sources}')
        for user in queryset:
            self.logi(f'USER = {user}')
            sps = list(user.sale_points.all())
            sps += list(get_model('refs.SalePoint').objects.filter(company__in=user.companies.values_list('id', flat=True)))
            self.logi(f'Sale Points = {len(sps)}')
            #self.logi(f'{sps}')
            for src in sources:
                for sp in sps:
                    ntfopt = None
                    try:
                        ntfopt = get_model('core.NotificationOption').objects.get(owner_id=user.id, source_id=src.id, sale_point_id=sp.id)
                    except ObjectDoesNotExist as e:
                        try:
                            ntfopt = get_model('core.NotificationOption')(owner_id=user.id, source_id=src.id, sale_point_id=sp.id)
                        except Exception as e:
                            self.loge(e)
                        else:
                            try:
                                ntfopt.save()
                            except Exception as e:
                                self.loge(e)
                    except Exception as e:
                        self.loge(e)
                    if ntfopt:
                        ntfopt.types.add(push_type)
                        changed += 1
        self.message_user(request, f'🆗 {changed}', messages.SUCCESS)
    set_all_push_notifications.short_description = f'🔔{_("set all push notifications")}'

    def queryset_to_xls(self, request, queryset, fields={}, exclude_fields=['password','favorites','notificationoption','notificationdelay','notificationtask','ingredientstoragehistory','logentry','gcmdevice','apnsdevice','wnsdevice','webpushdevice', 'avatar', 'groups', 'user_permissions', 'companies', 'sale_points']):
        import xlsxwriter
        output = None
        field_names = fields.keys() if fields else []
        if queryset.count():
            if not field_names:
                for field in queryset.model._meta.get_fields():
                    if field.name and field.name not in exclude_fields:
                        field_names.append(field.name)
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet()
            #date_format = workbook.add_format({'num_format': 'mmmm d yyyy'})
            #money_format = workbook.add_format({'num_format': '$#,##0'})
            caption_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'top'})
            worksheet.set_row(0, None, caption_format)
            default_format = workbook.add_format({'valign': 'top'})
            bool_format = workbook.add_format({'align': 'center', 'valign': 'top'})
            col = 0
            for field_name in field_names:
                col = self.worksheet_cell_write(worksheet, 0, col, fields.get(field_name, field_name))
            #worksheet.set_column(0, col-1, 20)
            text_wrap_format = workbook.add_format({'text_wrap': True, 'valign': 'top'})
            row = 1
            for item in queryset:
                col = 0
                for field_name in field_names:
                    if not hasattr(item, field_name):
                        col += 1
                        continue
                    else:
                        if not field_name:
                            col += 1
                            continue
                    try:
                        value = getattr(item, field_name)
                    except AttributeError as e:
                        col += 1
                        self.loge(e)
                    except Exception as e:
                        col += 1
                        self.loge(e)
                    else:
                        if value is None:
                            col += 1
                        else:
                            format_value = default_format
                            tvalue = None
                            if isinstance(value, datetime):
                                #tvalue = 'as_datetime'
                                #format_value = date_format
                                value = f'{value.strftime("%Y.%m.%d %H:%M:%S")}'
                            elif isinstance(value, bool):
                                format_value = bool_format
                                value = _('yes') if value else _('no')
                            elif isinstance(value, (int, float)):#value.isdigit():
                                #format_value = money_format
                                tvalue = 'as_number'
                            elif value.__class__.__name__ == 'ManyRelatedManager':
                                items = value.values_list('name', flat=True)
                                if items:
                                    format_value = text_wrap_format
                                    value = ';\n'.join([it for it in items])
                                    scount = value.count('\n') + 1
                                    if scount > 1:
                                        worksheet.set_row(row, scount*12)
                                else:
                                    value = ''
                            elif not isinstance(value, str):
                                value = f'{value}'
                            col = self.worksheet_cell_write(worksheet, row, col, value, tvalue, format_value)
                row += 1
            worksheet.autofit()
            workbook.close()
            output.seek(0)
        return output

    def selected_to_xls(self, request, queryset):
        output = self.queryset_to_xls(request, queryset)
        if output:
            fn = '{}.xlsx'.format(django_timezone.now().strftime('%Y%m%d%H%M%S'))
            self.message_user(request, f'🆗 {_("Finished")} ✏️({fn})', messages.SUCCESS)
            return FileResponse(output, as_attachment=True, filename=fn)
        self.message_user(request, _('please select items'), messages.ERROR)
    selected_to_xls.short_description = f'🏗️{_("selected to xls file")}'

    def filtered_by_selected_to_xls(self, request, queryset):
        if not queryset.count():
            self.message_user(request, _('please select items'), messages.ERROR)
        users = get_users_by_owner(queryset.first())
        ucount = users.count()
        self.logi('USERS COUNT', ucount)
        if ucount:
            output = self.queryset_to_xls(request, users, {'username': _('Login'), 'first_name': _('Name of people'), 'last_name': _('Family name'), 'last_login': _('Last of logged date'), 'is_active': _('Active'), 'companies': _('Companies'), 'sale_points': _('Objects')})
            if output:
                fn = '{}.xlsx'.format(django_timezone.now().strftime('%Y%m%d%H%M%S'))
                self.message_user(request, f'🆗 {_("Finished")} ✏️({fn})', messages.SUCCESS)
                return FileResponse(output, as_attachment=True, filename=fn)
            self.message_user(request, _('unknown error'), messages.ERROR)
        else:
            self.message_user(request, _('not found users'), messages.ERROR)
    filtered_by_selected_to_xls.short_description = f'🏗️{_("List of available by selected in xlsx")}'

    def add_notifies_by_last_from_same_notify_group_and_company(self, request, queryset):
        if not queryset.count():
            self.message_user(request, _('please select items'), messages.ERROR)
        else:
            msg, errors = 'CREATED COUNT = 0', ''
            created_items, owner_types = [], {}
            for u in queryset:
                u_notifies = get_model('core.NotificationOption').objects.filter(owner=u).order_by('id')
                last_notify = u_notifies.last()
                if not last_notify:
                    errmsg = f'{u} notifications empty'
                    errors = f'{errors}<br>{errmsg}' if errors else f'{errmsg}'
                    continue
                u_source_ids = []
                for nopt in u_notifies:
                    if nopt.source_id:
                        u_source_ids.append(nopt.source_id)
                if not last_notify.company_id:
                    errmsg = f'{u} {last_notify} company empty'
                    errors = f'{errors}<br>{errmsg}' if errors else f'{errmsg}'
                    continue
                if not last_notify.source.group_id:
                    errmsg = f'{u} {last_notify.source} group empty'
                    errors = f'{errors}<br>{errmsg}' if errors else f'{errmsg}'
                    continue
                owner_types[u.id] = last_notify.types.all()
                for ns in get_model('refs.NotificationSource').objects.filter(group=last_notify.source.group_id).exclude(id__in=u_source_ids):
                    created_items.append(get_model('core.NotificationOption')(owner=u, source=ns, company_id=last_notify.company_id))
            if created_items:
                try:
                    items = get_model('core.NotificationOption').objects.bulk_create(created_items)
                except Exception as e:
                    self.logw(e); errors = f'{errors}<br>{e}' if errors else f'{e}'
                else:
                    msg = f'CREATED COUNT = {len(items)}'
                    for it in items:
                        tps = owner_types.get(it.owner_id, [])
                        if tps:
                            it.types.add(*tps)
            if errors:
                msg += f'<br>{errors}'
            self.message_user(request, mark_safe(msg), messages.SUCCESS)
    add_notifies_by_last_from_same_notify_group_and_company.short_description = f'🔔{_("add notifies by last from same notify group and company")}'

admin.site.register(User, UserAdmin)
